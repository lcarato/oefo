"""
Excel workbook generator for OEFO observations.

Generates professional Excel workbooks with multiple sheets:
- Summary statistics by technology and country
- Full observation data in pivot-ready format
- WACC ranges by technology
- WACC ranges by country
- Methodology notes and documentation
"""

from pathlib import Path
from typing import Optional
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule


class ExcelOutputGenerator:
    """Generate Excel workbooks from OEFO observations."""

    def __init__(self):
        """Initialize the Excel output generator."""
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_font = Font(bold=True, size=14)
        self.center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_alignment = Alignment(horizontal='left', vertical='center')

    def generate_workbook(self, observations: pd.DataFrame, output_path: str) -> Path:
        """
        Generate a complete Excel workbook from observations.

        Args:
            observations: DataFrame containing observations with required columns
            output_path: Path where the workbook should be saved

        Returns:
            Path to the created workbook

        Raises:
            ValueError: If observations DataFrame is empty or missing required columns
        """
        if observations.empty:
            raise ValueError("Cannot generate workbook from empty observations DataFrame")

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        self.create_summary_sheet(wb, observations)
        self.create_data_sheet(wb, observations)
        self.create_provenance_sheet(wb, observations)
        self.create_technology_sheet(wb, observations)
        self.create_country_sheet(wb, observations)
        self.create_methodology_sheet(wb)

        # Apply formatting to all sheets
        self.apply_formatting(wb)

        # Save workbook
        wb.save(output_path)
        return output_path

    def create_summary_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Create summary statistics sheet with WACC by technology and country.

        Args:
            wb: Workbook to add sheet to
            df: Observations DataFrame
        """
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "OEFO Summary Statistics"
        title_cell.font = self.title_font
        title_cell.alignment = self.center_alignment

        # Timestamp
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Statistics
        ws['A4'] = "Overall Statistics"
        ws['A4'].font = Font(bold=True, size=12)

        stats_data = [
            ['Total Observations:', len(df)],
            ['Countries Covered:', df['country'].nunique() if 'country' in df.columns else 0],
            ['Technologies Covered:', df['technology'].nunique() if 'technology' in df.columns else 0],
            ['Date Range:', f"{df['date'].min()} to {df['date'].max()}" if 'date' in df.columns else "N/A"],
        ]

        for idx, (label, value) in enumerate(stats_data, 6):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value
            ws[f'B{idx}'].alignment = self.left_alignment

        # Technology × Country pivot summary
        if 'technology' in df.columns and 'country' in df.columns:
            pivot_start_row = 12
            ws[f'A{pivot_start_row}'] = "WACC by Technology × Country (Mean %)"
            ws[f'A{pivot_start_row}'].font = Font(bold=True, size=12)

            # Create pivot table
            wacc_col = 'wacc_percent' if 'wacc_percent' in df.columns else 'wacc'
            if wacc_col in df.columns:
                pivot = pd.pivot_table(
                    df,
                    values=wacc_col,
                    index='technology',
                    columns='country',
                    aggfunc='mean'
                )

                # Write pivot table
                for r_idx, (row_label, row_data) in enumerate(
                    [(None, pivot.columns)] + list(pivot.iterrows()),
                    pivot_start_row + 1
                ):
                    for c_idx, value in enumerate(
                        ([None] + list(row_data) if row_label is None else [row_label] + list(row_data)),
                        0
                    ):
                        cell = ws.cell(row=r_idx, column=c_idx + 1)
                        cell.value = value
                        if r_idx == pivot_start_row + 1:
                            cell.fill = self.header_fill
                            cell.font = self.header_font
                        elif c_idx == 0:
                            cell.font = Font(bold=True)
                        if isinstance(value, (int, float)) and not pd.isna(value):
                            cell.number_format = '0.00'
                        cell.border = self.thin_border
                        cell.alignment = self.center_alignment

    def create_data_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Create full data sheet with all observations in pivot-ready format.

        Args:
            wb: Workbook to add sheet to
            df: Observations DataFrame
        """
        ws = wb.create_sheet("Data")

        # Add title
        ws.merge_cells('A1:Z1')
        title_cell = ws['A1']
        title_cell.value = "Full Observation Data"
        title_cell.font = self.title_font
        title_cell.alignment = self.center_alignment

        # Write headers
        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = column
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
            cell.alignment = self.center_alignment

        # Write data
        for row_idx, (_, row) in enumerate(df.iterrows(), 3):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = self.thin_border
                cell.alignment = self.left_alignment

                # Format numeric columns
                if isinstance(value, float):
                    if 'percent' in df.columns[col_idx - 1].lower():
                        cell.number_format = '0.00'
                    else:
                        cell.number_format = '0.0000'

        # Auto-size columns
        self._auto_size_columns(ws)

        # Add data validation conditional formatting for confidence levels
        if 'confidence_level' in df.columns:
            conf_col = df.columns.get_loc('confidence_level') + 1
            for row in range(3, len(df) + 3):
                cell = ws.cell(row=row, column=conf_col)
                if cell.value == 'High':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif cell.value == 'Medium':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif cell.value == 'Low':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def create_provenance_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Create provenance/traceability sheet mapping each observation to its source.

        This sheet enables any user to trace every data point back to the exact
        page and quote in the original source document.

        Columns: observation_id, source_document_url, source_page_number,
                 source_quote, source_institution, source_type,
                 extraction_method, extraction_tier, confidence_level,
                 traceability_level

        Args:
            wb: Workbook to add sheet to
            df: Observations DataFrame
        """
        ws = wb.create_sheet("Provenance")

        # Title
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = "Data Provenance & Traceability"
        title_cell.font = self.title_font
        title_cell.alignment = self.center_alignment

        # Subtitle explaining purpose
        ws.merge_cells('A2:J2')
        ws['A2'] = (
            "Every row links an observation to its original source document, "
            "page number, and verbatim supporting quote."
        )
        ws['A2'].font = Font(italic=True, size=10)

        # Define provenance columns with preferred order
        provenance_cols = [
            'observation_id',
            'source_document_url',
            'source_page_number',
            'source_quote',
            'source_institution',
            'source_type',
            'source_document_id',
            'source_table_or_section',
            'extraction_method',
            'extraction_tier',
            'confidence_level',
            'traceability_level',
            'extraction_date',
        ]

        # Filter to columns that actually exist in the DataFrame
        available_cols = [c for c in provenance_cols if c in df.columns]

        # If key provenance columns are missing, still show what we have
        if not available_cols:
            available_cols = [c for c in df.columns if any(
                kw in c.lower() for kw in
                ['source', 'provenance', 'trace', 'extract', 'confidence', 'observation_id']
            )]

        if not available_cols:
            ws['A4'] = "No provenance columns found in the dataset."
            return

        # Write headers
        for col_idx, col_name in enumerate(available_cols, 1):
            cell = ws.cell(row=3, column=col_idx)
            # Make headers human-readable
            cell.value = col_name.replace('_', ' ').title()
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.thin_border
            cell.alignment = self.center_alignment

        # Write data
        for row_idx, (_, row) in enumerate(df.iterrows(), 4):
            for col_idx, col_name in enumerate(available_cols, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row.get(col_name, '')
                cell.value = value
                cell.border = self.thin_border
                cell.alignment = Alignment(
                    horizontal='left', vertical='top', wrap_text=True
                )

                # Colour-code traceability level
                if col_name == 'traceability_level':
                    if str(value).lower() == 'full':
                        cell.fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                        )
                    elif str(value).lower() == 'partial':
                        cell.fill = PatternFill(
                            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                        )
                    elif str(value).lower() == 'minimal':
                        cell.fill = PatternFill(
                            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                        )

                # Make URLs clickable
                if col_name == 'source_document_url' and value:
                    cell.hyperlink = str(value)
                    cell.font = Font(color="0563C1", underline="single")

        # Set column widths
        col_widths = {
            'observation_id': 20,
            'source_document_url': 50,
            'source_page_number': 12,
            'source_quote': 60,
            'source_institution': 20,
            'source_type': 18,
            'source_document_id': 20,
            'source_table_or_section': 25,
            'extraction_method': 15,
            'extraction_tier': 12,
            'confidence_level': 14,
            'traceability_level': 16,
            'extraction_date': 14,
        }
        for col_idx, col_name in enumerate(available_cols, 1):
            letter = ws.cell(row=3, column=col_idx).column_letter
            ws.column_dimensions[letter].width = col_widths.get(col_name, 15)

        # Freeze panes
        ws.freeze_panes = "A4"

    def create_technology_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Create technology analysis sheet with WACC ranges by technology.

        Args:
            wb: Workbook to add sheet to
            df: Observations DataFrame
        """
        ws = wb.create_sheet("By Technology")

        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "WACC Analysis by Technology"
        title_cell.font = self.title_font
        title_cell.alignment = self.center_alignment

        if 'technology' in df.columns:
            wacc_col = 'wacc_percent' if 'wacc_percent' in df.columns else 'wacc'
            if wacc_col in df.columns:
                tech_stats = df.groupby('technology')[wacc_col].agg([
                    'count', 'mean', 'std', 'min', 'max'
                ]).round(2)

                headers = ['Technology', 'Count', 'Mean', 'Std Dev', 'Min', 'Max']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col_idx)
                    cell.value = header
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.border = self.thin_border
                    cell.alignment = self.center_alignment

                for row_idx, (tech, row_data) in enumerate(tech_stats.iterrows(), 4):
                    ws.cell(row=row_idx, column=1).value = tech
                    for col_idx, value in enumerate(row_data, 2):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.border = self.thin_border
                        cell.alignment = self.center_alignment
                        if col_idx > 1:
                            cell.number_format = '0.00'

                self._auto_size_columns(ws)

    def create_country_sheet(self, wb: Workbook, df: pd.DataFrame) -> None:
        """
        Create country analysis sheet with WACC ranges by country.

        Args:
            wb: Workbook to add sheet to
            df: Observations DataFrame
        """
        ws = wb.create_sheet("By Country")

        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "WACC Analysis by Country"
        title_cell.font = self.title_font
        title_cell.alignment = self.center_alignment

        if 'country' in df.columns:
            wacc_col = 'wacc_percent' if 'wacc_percent' in df.columns else 'wacc'
            if wacc_col in df.columns:
                country_stats = df.groupby('country')[wacc_col].agg([
                    'count', 'mean', 'std', 'min', 'max'
                ]).round(2).sort_values('mean', ascending=False)

                headers = ['Country', 'Count', 'Mean', 'Std Dev', 'Min', 'Max']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col_idx)
                    cell.value = header
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                    cell.border = self.thin_border
                    cell.alignment = self.center_alignment

                for row_idx, (country, row_data) in enumerate(country_stats.iterrows(), 4):
                    ws.cell(row=row_idx, column=1).value = country
                    for col_idx, value in enumerate(row_data, 2):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.border = self.thin_border
                        cell.alignment = self.center_alignment
                        if col_idx > 1:
                            cell.number_format = '0.00'

                self._auto_size_columns(ws)

    def create_methodology_sheet(self, wb: Workbook) -> None:
        """
        Create methodology reference sheet with notes and documentation.

        Args:
            wb: Workbook to add sheet to
        """
        ws = wb.create_sheet("Methodology")

        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "OEFO Methodology"
        title_cell.font = self.title_font
        title_cell.alignment = self.center_alignment

        methodology_text = """
OVERVIEW
The Open Energy Finance Observatory (OEFO) collects and standardizes capital cost and financing data for renewable energy projects globally.

DATA SOURCES
- IFC (International Finance Corporation)
- EBRD (European Bank for Reconstruction and Development)
- GCF (Green Climate Fund)
- SEC (US Securities and Exchange Commission)
- ANEEL (Brazil)
- AER (Australia)
- Ofgem (UK)
- FERC (US Federal Energy Regulatory Commission)

WACC METHODOLOGY
Weighted Average Cost of Capital (WACC) combines debt and equity costs weighted by their proportion:
WACC = (E/V × Re) + (D/V × Rd × (1-Tc))
Where:
- E/V = proportion of equity
- D/V = proportion of debt
- Re = cost of equity
- Rd = cost of debt
- Tc = corporate tax rate

CONFIDENCE LEVELS
- High: Data from primary official sources with clear methodology
- Medium: Data from reputable secondary sources
- Low: Data from estimates or limited documentation

DATA TRACEABILITY
Every observation in this database carries full provenance metadata enabling end-to-end traceability:

Traceability Levels:
- FULL: source_document_url + source_page_number + source_quote all present
- PARTIAL: some provenance fields missing (flagged for review)
- MINIMAL: only source_type and institution known (requires human verification)

Provenance fields per observation:
- source_document_url: URL where the original document was obtained
- source_document_id: Internal ID linking to the cached document
- source_page_number: Exact page number in the PDF
- source_quote: Verbatim text from the document supporting the value
- source_table_or_section: Table name or section heading if applicable
- extraction_tier: Which extraction method was used (Text/OCR/Vision/Human)
- extraction_method: Specific tool or model used

See the "Provenance" sheet for the complete traceability record.

CURRENCY
All values are in local currency unless otherwise specified. USD conversions available upon request.

DATA QUALITY
- All observations undergo 3-layer automated quality checks (rules, statistics, LLM)
- Traceability completeness is validated as part of QC Layer 1 (rule-based)
- Observations with MINIMAL traceability are automatically flagged for human review
- See QC report for validation results
- Missing data points are flagged and documented
"""

        row = 3
        for line in methodology_text.strip().split('\n'):
            ws[f'A{row}'] = line
            ws[f'A{row}'].alignment = self.left_alignment
            if line and not line.startswith(' '):
                ws[f'A{row}'].font = Font(bold=True, size=11)
            row += 1

        ws.column_dimensions['A'].width = 100

    def apply_formatting(self, wb: Workbook) -> None:
        """
        Apply professional formatting to all sheets in the workbook.

        Args:
            wb: Workbook to format
        """
        for ws in wb.sheetnames:
            sheet = wb[ws]

            # Set default column width
            from openpyxl.cell.cell import MergedCell
            for column in sheet.columns:
                first_cell = column[0]
                if isinstance(first_cell, MergedCell):
                    continue
                if first_cell.column_letter != 'A':
                    sheet.column_dimensions[first_cell.column_letter].width = 14

            # Freeze panes on data sheet
            if ws == "Data":
                sheet.freeze_panes = "A3"
            elif ws == "Summary":
                sheet.freeze_panes = "A4"

    def _auto_size_columns(self, ws) -> None:
        """
        Auto-size columns based on content.

        Args:
            ws: Worksheet to resize
        """
        from openpyxl.cell.cell import MergedCell
        for column in ws.columns:
            max_length = 0
            # Skip merged cells which don't have column_letter
            first_cell = column[0]
            if isinstance(first_cell, MergedCell):
                continue
            column_letter = first_cell.column_letter

            for cell in column:
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
